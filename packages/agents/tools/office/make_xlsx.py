#!/usr/bin/env python3
"""Build a spreadsheet from a JSON spec. No API to learn, one command:

    make-sheet spec.json out.xlsx

spec.json:

    {"sheets": [
       {"name": "Q3", "columns": ["Item", "Cost"],
        "rows": [["Hosting", 120], ["Backups", 40]],
        "widths": [28, 12]}
     ]}

columns is the header row, drawn bold. widths is optional. Numbers go in as
JSON numbers, not strings, or the reader's own formulas will not add.

This exists because a model writing openpyxl code from memory invents
imports and APIs; a model filling in this spec and running one command does
not. The spec is the whole surface.
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from artifact_path import resolve_output  # noqa: E402


def build(spec: dict, out: Path) -> int:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    sheets = spec.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        raise SystemExit('spec has no "sheets": expected a non-empty list of {"name","columns","rows"}')
    wb = Workbook()
    wb.remove(wb.active)
    for i, s in enumerate(sheets):
        if not isinstance(s, dict):
            raise SystemExit(f'sheets[{i}] is not an object')
        ws = wb.create_sheet(title=str(s.get("name", f"Sheet{i + 1}")))
        cols = s.get("columns", [])
        if cols:
            ws.append(cols)
            for c in ws[1]:
                c.font = Font(bold=True)
        rows = s.get("rows", [])
        if not isinstance(rows, list):
            raise SystemExit(f'sheets[{i}].rows is not a list of rows')
        for row in rows:
            ws.append(row)
        for j, w in enumerate(s.get("widths", []), start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
    parent = os.path.dirname(str(out))
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(str(out))
    print(f"wrote {out} ({len(sheets)} sheets)")
    return 0


def main(argv: list[str]) -> int:
    # See make_docx.py: the one real cause of a wrong count is an unquoted
    # space in the output path, so the error says that instead of the usage.
    if len(argv) > 3:
        raise SystemExit(
            f"got {len(argv) - 1} arguments, expected 2: make-sheet spec.json out.xlsx —"
            " a path with a space must be quoted; better, name the file with dashes,"
            " like /artifacts/sheets/q3-budget.xlsx"
        )
    if len(argv) < 3:
        print(__doc__)
        return 2
    spec_path, out = Path(argv[1]), Path(resolve_output(argv[2]))
    try:
        spec = json.loads(spec_path.read_text("utf8"))
    except FileNotFoundError:
        raise SystemExit(f"{spec_path}: no such file — write the spec first, then run make-sheet")
    except json.JSONDecodeError as e:
        raise SystemExit(f"{spec_path} is not valid JSON: {e}")
    if not isinstance(spec, dict):
        raise SystemExit(f"{spec_path}: expected a JSON object, got {type(spec).__name__}")
    return build(spec, out)


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
