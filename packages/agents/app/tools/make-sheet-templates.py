# Three sheet templates, written from an empty workbook out.
#
# These replace three Vertex42 downloads whose embedded licence forbids
# publishing to a template gallery. Nothing here is copied from them beyond
# the IDEA each name already carries: a Gantt chart is tasks against weeks, an
# appointment schedule is a day of time slots, a savings calculator is
# compound interest down a column. Layout, palette, formulas and wording are
# this file's own.
#
# Formulas are written as formulas — Excel and LibreOffice compute them on
# open, and the engine's thumbnail conversion runs LibreOffice, so the cards
# show computed values without this script ever computing one.
#
# Run it where openpyxl exists; this deployment has no host pip, so:
#   sudo docker run --rm -v "$PWD":/work agents-search:1 sh -c \
#     "python3 -m pip install -q --break-system-packages openpyxl && \
#      python3 /work/make-sheet-templates.py"
# then POST each file as a template row + template file (see git history of
# this commit for the exact bodies).

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# One palette across the three, so they read as a set: slate ink, one accent
# per workbook, a pale grid.
INK = "1F2430"
GRID = "E7E7EC"
PALE = "F4F4F6"

def head_style(ws, cells, accent):
    for c in cells:
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(horizontal="center", vertical="center")

def title(ws, text, accent, span):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    t = ws["A1"]
    t.value = text
    t.font = Font(bold=True, size=16, color=INK)
    ws.row_dimensions[1].height = 28

thin = Side(style="thin", color=GRID)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

# --- 1. Gantt chart ----------------------------------------------------------
ACCENT_G = "2563EB"
wb = Workbook()
ws = wb.active
ws.title = "Gantt"
ws.sheet_view.showGridLines = False
WEEKS = 16
title(ws, "Project plan", ACCENT_G, 6 + WEEKS)

ws["A2"] = "Project start"
ws["A2"].font = Font(bold=True, color=INK)
ws["B2"] = "=TODAY()"
ws["B2"].number_format = "yyyy-mm-dd"

heads = ["Task", "Owner", "Start", "End", "Days", "Done"]
for i, h in enumerate(heads, start=1):
    ws.cell(row=4, column=i, value=h)
head_style(ws, [ws.cell(row=4, column=i) for i in range(1, 7)], ACCENT_G)
for w in range(WEEKS):
    c = ws.cell(row=4, column=7 + w, value=f"W{w + 1}")
    c.font = Font(bold=True, size=9, color=INK)
    c.alignment = Alignment(horizontal="center")
    ws.column_dimensions[get_column_letter(7 + w)].width = 3.4

widths = {"A": 30, "B": 14, "C": 12, "D": 12, "E": 7, "F": 7}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

rows = [
    ("Kick-off and scope", "PM", 0, 5),
    ("Research", "Ana", 3, 12),
    ("Design", "Ben", 10, 24),
    ("Build", "Team", 21, 45),
    ("Review", "PM", 42, 52),
    ("Launch", "Team", 52, 60),
]
for r, (task, owner, s, e) in enumerate(rows, start=5):
    ws.cell(row=r, column=1, value=task)
    ws.cell(row=r, column=2, value=owner)
    ws.cell(row=r, column=3, value=f"=$B$2+{s}")
    ws.cell(row=r, column=4, value=f"=$B$2+{e}")
    ws.cell(row=r, column=5, value=f"=D{r}-C{r}+1")
    ws.cell(row=r, column=6, value=0)
    ws.cell(row=r, column=6).number_format = "0%"
    ws.cell(row=r, column=3).number_format = "mm-dd"
    ws.cell(row=r, column=4).number_format = "mm-dd"
    for col in range(1, 7):
        ws.cell(row=r, column=col).border = box
    for w in range(WEEKS):
        ws.cell(row=r, column=7 + w).border = box

# The bars: a week column is filled when the task overlaps that week. The
# rule lives on the week area, so adding a task row above row 20 keeps
# working without touching the formatting.
first, last = 5, 20
area = f"G{first}:{get_column_letter(6 + WEEKS)}{last}"
overlap = (
    'AND($A5<>"",'
    "$B$2+(COLUMN()-7)*7<=$D5,"
    "$B$2+(COLUMN()-6)*7-1>=$C5)"
)
ws.conditional_formatting.add(
    area,
    FormulaRule(formula=[overlap], fill=PatternFill("solid", fgColor=ACCENT_G)),
)
ws["A22"] = "Add rows above row 20; the bars follow Start and End on their own."
ws["A22"].font = Font(italic=True, size=10, color="6B6B76")
wb.save("/work/gantt-chart.xlsx")

# --- 2. Appointment schedule -------------------------------------------------
ACCENT_A = "0F766E"
wb = Workbook()
ws = wb.active
ws.title = "Week"
ws.sheet_view.showGridLines = False
DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
title(ws, "Appointment schedule", ACCENT_A, 2 + len(DAYS))

ws["A2"] = "Week of"
ws["A2"].font = Font(bold=True, color=INK)
ws["B2"] = "=TODAY()-WEEKDAY(TODAY(),3)"
ws["B2"].number_format = "yyyy-mm-dd"
ws["D2"] = "Slot minutes"
ws["D2"].font = Font(bold=True, color=INK)
ws["E2"] = 30

ws["A4"] = "Time"
head_style(ws, [ws["A4"]], ACCENT_A)
for i, d in enumerate(DAYS):
    c = ws.cell(row=4, column=2 + i, value=d)
    head_style(ws, [c], ACCENT_A)
    ws.column_dimensions[get_column_letter(2 + i)].width = 22
ws.column_dimensions["A"].width = 10

SLOTS = 20  # 09:00 onwards, one row per slot
for i in range(SLOTS):
    r = 5 + i
    t = ws.cell(row=r, column=1, value=f"=TIME(9,0,0)+$E$2*{i}/1440")
    t.number_format = "hh:mm"
    t.font = Font(bold=True, size=10, color=INK)
    t.alignment = Alignment(horizontal="center")
    for c in range(1, 2 + len(DAYS)):
        ws.cell(row=r, column=c).border = box
    if i % 2 == 0:
        for c in range(2, 2 + len(DAYS)):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=PALE)

ws.cell(row=5 + SLOTS + 1, column=1, value="Type a name into a slot to book it. Change slot minutes in E2 and every time follows.")
ws.cell(row=5 + SLOTS + 1, column=1).font = Font(italic=True, size=10, color="6B6B76")
wb.save("/work/appointment-schedule.xlsx")

# --- 3. Savings interest calculator ------------------------------------------
ACCENT_S = "B45309"
wb = Workbook()
ws = wb.active
ws.title = "Savings"
ws.sheet_view.showGridLines = False
title(ws, "Savings with compound interest", ACCENT_S, 6)

inputs = [
    ("Starting balance", 1000, "#,##0.00"),
    ("Monthly deposit", 100, "#,##0.00"),
    ("Yearly rate", 0.04, "0.00%"),
    ("Years", 15, "0"),
]
for i, (label, v, fmt) in enumerate(inputs):
    ws.cell(row=3 + i, column=1, value=label).font = Font(bold=True, color=INK)
    c = ws.cell(row=3 + i, column=2, value=v)
    c.number_format = fmt
    c.fill = PatternFill("solid", fgColor=PALE)
    c.border = box

for i, h in enumerate(["Year", "Deposited", "Interest", "Balance"], start=1):
    ws.cell(row=9, column=i, value=h)
head_style(ws, [ws.cell(row=9, column=i) for i in range(1, 5)], ACCENT_S)
for col, w in {"A": 18, "B": 14, "C": 14, "D": 14}.items():
    ws.column_dimensions[col].width = w

YEARS_ROWS = 40
for i in range(YEARS_ROWS):
    r = 10 + i
    ws.cell(row=r, column=1, value=f'=IF({i}<=$B$6,{i},"")')
    if i == 0:
        ws.cell(row=r, column=2, value="=$B$3")
        ws.cell(row=r, column=4, value="=$B$3")
        ws.cell(row=r, column=3, value=0)
    else:
        # Monthly compounding, deposits at month end — one row per year.
        ws.cell(row=r, column=2, value=f'=IF(A{r}="","",B{r-1}+$B$4*12)')
        ws.cell(row=r, column=4, value=(
            f'=IF(A{r}="","",'
            f"D{r-1}*(1+$B$5/12)^12"
            f"+$B$4*(((1+$B$5/12)^12-1)/($B$5/12)))"
        ))
        ws.cell(row=r, column=3, value=f'=IF(A{r}="","",D{r}-B{r})')
    for c in range(2, 5):
        ws.cell(row=r, column=c).number_format = "#,##0.00"
        ws.cell(row=r, column=c).border = box
    ws.cell(row=r, column=1).border = box

chart = LineChart()
chart.title = "Balance and deposits"
chart.height = 8
chart.width = 16
data = Reference(ws, min_col=2, max_col=4, min_row=9, max_row=9 + YEARS_ROWS)
cats = Reference(ws, min_col=1, min_row=10, max_row=9 + YEARS_ROWS)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
ws.add_chart(chart, "F3")
wb.save("/work/savings-interest.xlsx")
print("wrote 3 workbooks")
