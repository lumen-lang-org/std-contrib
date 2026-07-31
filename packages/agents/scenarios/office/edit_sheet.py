# Fill a workbook that already exists.
#
# The template's formulas are the reason to open it rather than build beside
# it: the variance and total columns compute, and they keep computing after
# the reader edits a cell. Writing rows into the existing sheet preserves
# them; generating a new workbook throws them away.
import os
from openpyxl import load_workbook


def fill(path, rows, sheet=None, start_row=2, first_col=1):
    """Write `rows` into the template's sheet, leaving its formulas alone.

    rows = [["Hosting", 120, 118], ...] — as many columns as the template
    expects a human to fill, no more: a row that reaches into the variance
    column overwrites the formula it was there to use.
    """
    wb = load_workbook(path)
    ws = wb[sheet] if sheet else wb.active
    for r, values in enumerate(rows, start=start_row):
        for c, value in enumerate(values, start=first_col):
            ws.cell(r, c, value)
    # Anything the template pre-filled below the data stays: the total row
    # and its SUMs are part of the document, not of this call.
    wb.save(path)
    print(f"wrote {len(rows)} rows into {os.path.basename(path)}")
    return len(rows)
