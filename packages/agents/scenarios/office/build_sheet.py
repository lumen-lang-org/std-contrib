# The make-sheet builder.
#
# spec = {
#   "path": "/artifacts/sheets/budget.xlsx",
#   "sheets": [
#     { "name": "Q3", "columns": ["Item", "Cost"],   # header row, bold
#       "rows": [["Hosting", 120], ["Backups", 40]],
#       "widths": [28, 12] },                        # optional column widths
#   ],
# }
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def build(spec):
    wb = Workbook()
    wb.remove(wb.active)
    for s in spec.get("sheets", []):
        ws = wb.create_sheet(title=s.get("name", "Sheet"))
        cols = s.get("columns", [])
        if cols:
            ws.append(cols)
            for c in ws[1]:
                c.font = Font(bold=True)
        for row in s.get("rows", []):
            ws.append(row)
        for i, w in enumerate(s.get("widths", []), start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
    path = spec["path"]
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print("wrote", path)
